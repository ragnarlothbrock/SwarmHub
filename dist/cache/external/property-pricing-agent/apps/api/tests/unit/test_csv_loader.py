"""
Comprehensive unit tests for data/csv_loader.py.

Covers DataLoaderCsv, DataLoaderExcel, and static helper methods:
- Initialization with Path, URL, str, None
- CSV and Excel loading with various encodings and error paths
- URL validation and GitHub URL conversion
- DataFrame formatting, column normalization, and data enrichment
- Bathroom/price_media fake generators
- camel_to_snake conversion
- DataLoaderExcel: sheet detection, load_df, detect_source_type
- Edge cases: missing files, malformed data, encoding issues, empty DataFrames
"""

import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import numpy as np
import pandas as pd
import pytest
from yarl import URL

from data.csv_loader import DataLoaderCsv, DataLoaderExcel


# ---------------------------------------------------------------------------
# DataLoaderCsv.__init__
# ---------------------------------------------------------------------------


class TestDataLoaderCsvInit:
    """Tests for DataLoaderCsv initialization."""

    def test_init_with_none(self):
        loader = DataLoaderCsv(None)
        assert loader.csv_path is None

    def test_init_with_valid_csv_path(self):
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
            f.write(b"city,price\nWarsaw,1000\n")
            path = Path(f.name)
        try:
            loader = DataLoaderCsv(path)
            assert loader.csv_path == path
        finally:
            path.unlink()

    def test_init_with_nonexistent_path_logs_warning(self, caplog):
        """Non-existent Path sets csv_path to None with a warning."""
        nonexistent = Path("/tmp/this_file_does_not_exist_abc123.csv")
        with caplog.at_level("WARNING"):
            loader = DataLoaderCsv(nonexistent)
        assert loader.csv_path is None
        assert "does not exists" in caplog.text

    def test_init_with_string_path(self):
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
            f.write(b"city,price\nKrakow,900\n")
            path = Path(f.name)
        try:
            loader = DataLoaderCsv(str(path))
            assert loader.csv_path == str(path)
        finally:
            path.unlink()

    def test_init_with_url_invalid_logs_warning(self, caplog):
        """Invalid URL sets csv_path to None with a warning."""
        bad_url = URL("not-a-valid-url")
        with caplog.at_level("WARNING"):
            loader = DataLoaderCsv(bad_url)
        assert loader.csv_path is None

    @patch("data.csv_loader.requests.head")
    def test_init_with_valid_url(self, mock_head):
        mock_head.return_value = MagicMock(status_code=200)
        url = URL("https://example.com/data.csv")
        loader = DataLoaderCsv(url)
        assert loader.csv_path == url


# ---------------------------------------------------------------------------
# DataLoaderCsv.url_exists
# ---------------------------------------------------------------------------


class TestUrlExists:
    """Tests for DataLoaderCsv.url_exists static method."""

    @patch("data.csv_loader.requests.head")
    def test_valid_url_returns_true(self, mock_head):
        mock_head.return_value = MagicMock(status_code=200)
        assert DataLoaderCsv.url_exists(URL("https://example.com/data.csv")) is True

    @patch("data.csv_loader.requests.head")
    def test_redirect_301_returns_true(self, mock_head):
        mock_head.return_value = MagicMock(status_code=301)
        assert DataLoaderCsv.url_exists(URL("https://example.com/old")) is True

    @patch("data.csv_loader.requests.head")
    def test_404_returns_false(self, mock_head):
        mock_head.return_value = MagicMock(status_code=404)
        assert DataLoaderCsv.url_exists(URL("https://example.com/missing")) is False

    def test_invalid_url_structure_returns_false(self):
        assert DataLoaderCsv.url_exists(URL("not-a-url")) is False

    @patch("data.csv_loader.requests.head")
    def test_request_exception_returns_false(self, mock_head):
        import requests

        mock_head.side_effect = requests.RequestException("timeout")
        assert DataLoaderCsv.url_exists(URL("https://example.com/data.csv")) is False


# ---------------------------------------------------------------------------
# DataLoaderCsv.convert_github_url_to_raw
# ---------------------------------------------------------------------------


class TestConvertGithubUrlToRaw:
    """Tests for convert_github_url_to_raw static method."""

    def test_github_blob_url_converted(self):
        url = "https://github.com/user/repo/blob/main/data.csv"
        result = DataLoaderCsv.convert_github_url_to_raw(url)
        assert result == "https://raw.githubusercontent.com/user/repo/main/data.csv"

    def test_non_github_url_unchanged(self):
        url = "https://example.com/data.csv"
        result = DataLoaderCsv.convert_github_url_to_raw(url)
        assert result == url

    def test_github_url_without_blob_unchanged(self):
        url = "https://github.com/user/repo/tree/main/data"
        result = DataLoaderCsv.convert_github_url_to_raw(url)
        assert result == url

    def test_empty_string_unchanged(self):
        assert DataLoaderCsv.convert_github_url_to_raw("") == ""

    def test_raw_url_already_unchanged(self):
        url = "https://raw.githubusercontent.com/user/repo/main/data.csv"
        result = DataLoaderCsv.convert_github_url_to_raw(url)
        assert result == url


# ---------------------------------------------------------------------------
# DataLoaderCsv.load_df
# ---------------------------------------------------------------------------


class TestLoadDf:
    """Tests for DataLoaderCsv.load_df."""

    def test_load_csv_from_path(self):
        df_in = pd.DataFrame({"city": ["Krakow"], "price": [900], "rooms": [2]})
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "data.csv"
            df_in.to_csv(p, index=False)
            loader = DataLoaderCsv(p)
            df_out = loader.load_df()
        assert list(df_out.columns) == ["city", "price", "rooms"]
        assert len(df_out) == 1

    def test_load_df_none_path_raises_value_error(self):
        loader = DataLoaderCsv(None)
        with pytest.raises(ValueError, match="No CSV/Excel path provided"):
            loader.load_df()

    @patch("data.csv_loader.pd.read_csv")
    def test_load_df_csv_fallback_to_utf8(self, mock_read_csv):
        """First read_csv fails, fallback with utf-8 engine='python' succeeds."""
        df_expected = pd.DataFrame({"city": ["Berlin"], "price": [1200]})
        mock_read_csv.side_effect = [
            Exception("initial failure"),
            df_expected,
        ]

        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "data.csv"
            p.write_text("city,price\nBerlin,1200\n")
            loader = DataLoaderCsv(str(p))

        df_out = loader.load_df()
        assert len(df_out) == 1

    @patch("data.csv_loader.pd.read_csv")
    def test_load_df_csv_fallback_to_latin1(self, mock_read_csv):
        """First two read_csv calls fail, latin-1 fallback succeeds."""
        df_expected = pd.DataFrame({"city": ["Munchen"], "price": [1100]})
        mock_read_csv.side_effect = [
            Exception("initial failure"),
            Exception("utf-8 failure"),
            df_expected,
        ]

        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "data.csv"
            p.write_text("city,price\nMunchen,1100\n")
            loader = DataLoaderCsv(str(p))

        df_out = loader.load_df()
        assert len(df_out) == 1

    @patch("data.csv_loader.pd.read_csv")
    def test_load_df_csv_all_attempts_fail_raises(self, mock_read_csv):
        """All read_csv attempts fail: raises with combined error message."""
        mock_read_csv.side_effect = Exception("fail1")
        loader = DataLoaderCsv("nonexistent.csv")
        # The last call is with latin-1 which also fails
        mock_read_csv.side_effect = [
            Exception("fail1"),
            Exception("fail2"),
            Exception("fail3"),
        ]
        with pytest.raises(Exception, match="Failed to load CSV"):
            loader.load_df()

    def test_load_excel_from_path(self):
        pytest.importorskip("openpyxl")
        df_in = pd.DataFrame({"city": ["Warsaw"], "price": [1200], "rooms": [3]})
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "data.xlsx"
            df_in.to_excel(p, index=False)
            loader = DataLoaderCsv(p)
            df_out = loader.load_df()
        assert list(df_out.columns) == ["city", "price", "rooms"]
        assert len(df_out) == 1

    def test_load_df_url_csv(self):
        """String URL is treated as a CSV source."""
        df_expected = pd.DataFrame({"a": [1]})

        with patch("data.csv_loader.pd.read_csv", return_value=df_expected):
            loader = DataLoaderCsv("https://example.com/data.csv")
            loader.csv_path = "https://example.com/data.csv"  # Force path
            df_out = loader.load_df()

        assert len(df_out) == 1

    @patch("data.csv_loader.pd.read_excel")
    def test_load_df_excel_xls_engine(self, mock_read_excel):
        mock_read_excel.return_value = pd.DataFrame({"a": [1]})
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "data.xls"
            p.write_text("dummy")
            loader = DataLoaderCsv(str(p))
            df_out = loader.load_df()
        assert len(df_out) == 1

    @patch("data.csv_loader.pd.read_excel")
    def test_load_df_excel_xlsx_fallback_on_value_error(self, mock_read_excel):
        """read_excel with engine='openpyxl' raises ValueError -> fallback without engine."""
        mock_read_excel.side_effect = [ValueError("engine error"), pd.DataFrame({"a": [1]})]
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "data.xlsx"
            p.write_text("dummy")
            loader = DataLoaderCsv(str(p))
            df_out = loader.load_df()
        assert len(df_out) == 1

    @patch("data.csv_loader.pd.read_excel")
    def test_load_df_excel_import_error_raises(self, mock_read_excel):
        mock_read_excel.side_effect = ImportError("no openpyxl")
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "data.xlsx"
            p.write_text("dummy")
            loader = DataLoaderCsv(str(p))
        with pytest.raises(ImportError, match="Excel input requires optional"):
            loader.load_df()

    @patch("data.csv_loader.pd.read_excel")
    def test_load_df_excel_generic_exception_raises(self, mock_read_excel):
        mock_read_excel.side_effect = Exception("corrupt file")
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "data.xlsx"
            p.write_text("dummy")
            loader = DataLoaderCsv(str(p))
        with pytest.raises(Exception, match="Failed to load Excel file"):
            loader.load_df()

    @patch("data.csv_loader.pd.read_excel")
    def test_load_df_excel_generic_exception_xls(self, mock_read_excel):
        mock_read_excel.side_effect = Exception("corrupt xls")
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "data.xls"
            p.write_text("dummy")
            loader = DataLoaderCsv(str(p))
        with pytest.raises(Exception, match="Failed to load Excel file"):
            loader.load_df()


# ---------------------------------------------------------------------------
# DataLoaderCsv.load_format_df
# ---------------------------------------------------------------------------


class TestLoadFormatDf:
    """Tests for DataLoaderCsv.load_format_df."""

    def test_load_format_df_returns_formatted(self):
        df = pd.DataFrame({"city": ["Warsaw"], "price": [1000], "rooms": [2]})
        loader = DataLoaderCsv(None)
        result = loader.load_format_df(df)
        assert "city" in result.columns
        assert "price" in result.columns

    def test_load_format_df_with_rows_count(self):
        df = pd.DataFrame({"city": ["Warsaw", "Krakow", "Gdansk"], "price": [1000, 900, 800]})
        loader = DataLoaderCsv(None)
        result = loader.load_format_df(df, rows_count=2)
        assert len(result) == 2


# ---------------------------------------------------------------------------
# DataLoaderCsv.bathrooms_fake
# ---------------------------------------------------------------------------


class TestBathroomsFake:
    """Tests for bathrooms_fake static method."""

    def test_nan_returns_one(self):
        assert DataLoaderCsv.bathrooms_fake(float("nan")) == 1.0

    def test_less_than_two_rooms_returns_one(self):
        assert DataLoaderCsv.bathrooms_fake(1.0) == 1.0
        assert DataLoaderCsv.bathrooms_fake(0.5) == 1.0
        assert DataLoaderCsv.bathrooms_fake(1.9) == 1.0

    def test_two_or_more_rooms_returns_valid(self):
        v = DataLoaderCsv.bathrooms_fake(2.0)
        assert v in (1.0, 2.0)

    def test_many_rooms_returns_valid(self):
        v = DataLoaderCsv.bathrooms_fake(5.0)
        assert v in (1.0, 2.0)


# ---------------------------------------------------------------------------
# DataLoaderCsv.price_media_fake
# ---------------------------------------------------------------------------


class TestPriceMediaFake:
    """Tests for price_media_fake static method."""

    def test_returns_non_negative(self):
        result = DataLoaderCsv.price_media_fake(1000.0)
        assert result >= 0

    def test_within_20_percent_of_price(self):
        price = 1000.0
        for _ in range(50):
            result = DataLoaderCsv.price_media_fake(price)
            assert result <= 0.2 * price

    def test_zero_price_returns_zero(self):
        result = DataLoaderCsv.price_media_fake(0.0)
        assert result == 0.0

    def test_rounded_to_two_decimals(self):
        result = DataLoaderCsv.price_media_fake(999.99)
        assert result == round(result, 2)


# ---------------------------------------------------------------------------
# DataLoaderCsv.camel_to_snake
# ---------------------------------------------------------------------------


class TestCamelToSnake:
    """Tests for camel_to_snake static method."""

    def test_simple_camel_case(self):
        assert DataLoaderCsv.camel_to_snake("camelCase") == "camel_case"

    def test_pascal_case(self):
        assert DataLoaderCsv.camel_to_snake("PascalCase") == "pascal_case"

    def test_already_snake_case(self):
        assert DataLoaderCsv.camel_to_snake("already_snake") == "already_snake"

    def test_single_word_lowercase(self):
        assert DataLoaderCsv.camel_to_snake("price") == "price"

    def test_single_word_uppercase(self):
        assert DataLoaderCsv.camel_to_snake("PRICE") == "price"

    def test_with_numbers(self):
        assert DataLoaderCsv.camel_to_snake("area2Sqm") == "area2_sqm"

    def test_consecutive_uppercase(self):
        assert DataLoaderCsv.camel_to_snake("hasHTTP") == "has_http"

    def test_empty_string(self):
        assert DataLoaderCsv.camel_to_snake("") == ""


# ---------------------------------------------------------------------------
# DataLoaderCsv.format_df
# ---------------------------------------------------------------------------


class TestFormatDf:
    """Comprehensive tests for format_df static method."""

    def test_basic_formatting_adds_missing_columns(self):
        df = pd.DataFrame(
            {
                "City": ["Krakow", None],
                "Rooms": [2, None],
                "Price": [900, 1200],
                "has_parking": ["yes", None],
                "has_balcony": [1, 0],
            }
        )
        out = DataLoaderCsv.format_df(df)
        assert "city" in out.columns
        assert "rooms" in out.columns
        assert "bathrooms" in out.columns
        assert isinstance(out.loc[0, "rooms"], float)
        assert len(out) == 2

    def test_columns_converted_to_snake_case(self):
        df = pd.DataFrame({"CamelCase": [1], "AnotherCol": [2]})
        out = DataLoaderCsv.format_df(df)
        assert "camel_case" in out.columns
        assert "another_col" in out.columns

    def test_price_synonym_renamed(self):
        """Columns with 'cost' or 'rent' in the name are renamed to 'price'."""
        df = pd.DataFrame({"cost": [500], "rooms": [2]})
        out = DataLoaderCsv.format_df(df)
        assert "price" in out.columns
        assert out.loc[0, "price"] == 500

    def test_rent_synonym_renamed_to_price(self):
        df = pd.DataFrame({"rent_amount": [800], "rooms": [2]})
        out = DataLoaderCsv.format_df(df)
        assert "price" in out.columns

    def test_city_synonym_renamed(self):
        df = pd.DataFrame({"location": ["Berlin"], "price": [1200]})
        out = DataLoaderCsv.format_df(df)
        assert "city" in out.columns

    def test_town_synonym_renamed_to_city(self):
        df = pd.DataFrame({"town": ["Poznan"], "price": [900]})
        out = DataLoaderCsv.format_df(df)
        assert "city" in out.columns

    def test_place_synonym_renamed_to_city(self):
        df = pd.DataFrame({"place": ["Gdansk"], "price": [700]})
        out = DataLoaderCsv.format_df(df)
        assert "city" in out.columns

    def test_no_city_column_adds_unknown(self):
        df = pd.DataFrame({"price": [1000], "rooms": [2]})
        out = DataLoaderCsv.format_df(df)
        assert "city" in out.columns
        assert out.loc[0, "city"] == "Unknown"

    def test_rooms_synonym_renamed(self):
        df = pd.DataFrame({"city": ["Warsaw"], "price": [1000], "bedroom": [3]})
        out = DataLoaderCsv.format_df(df)
        assert "rooms" in out.columns

    def test_no_rooms_column_adds_default(self):
        df = pd.DataFrame({"city": ["Warsaw"], "price": [1000]})
        out = DataLoaderCsv.format_df(df)
        assert "rooms" in out.columns
        assert out.loc[0, "rooms"] == 2.0

    def test_rooms_nan_filled_with_default(self):
        df = pd.DataFrame({"city": ["Warsaw"], "price": [1000], "rooms": [None]})
        out = DataLoaderCsv.format_df(df)
        assert out.loc[0, "rooms"] == 2.0

    def test_area_synonym_renamed(self):
        df = pd.DataFrame({"city": ["Warsaw"], "price": [1000], "size": [50]})
        out = DataLoaderCsv.format_df(df)
        assert "area_sqm" in out.columns

    def test_sqm_synonym_renamed_to_area(self):
        df = pd.DataFrame({"city": ["Warsaw"], "price": [1000], "sqm": [45]})
        out = DataLoaderCsv.format_df(df)
        assert "area_sqm" in out.columns

    def test_square_synonym_renamed_to_area(self):
        df = pd.DataFrame({"city": ["Warsaw"], "price": [1000], "square_meters": [60]})
        out = DataLoaderCsv.format_df(df)
        assert "area_sqm" in out.columns

    def test_currency_synonym_renamed(self):
        df = pd.DataFrame({"city": ["Berlin"], "price": [1200], "curr": ["EUR"]})
        out = DataLoaderCsv.format_df(df)
        assert "currency" in out.columns

    def test_currency_default_for_polish_cities(self):
        df = pd.DataFrame({"city": ["Warsaw"], "price": [1000]})
        out = DataLoaderCsv.format_df(df)
        assert out.loc[0, "currency"] == "PLN"

    def test_currency_default_unknown_for_non_polish(self):
        df = pd.DataFrame({"city": ["Berlin"], "price": [1000]})
        out = DataLoaderCsv.format_df(df)
        assert out.loc[0, "currency"] == "Unknown"

    def test_listing_type_default_rent(self):
        df = pd.DataFrame({"city": ["Warsaw"], "price": [5000], "rooms": [2]})
        out = DataLoaderCsv.format_df(df)
        assert out.loc[0, "listing_type"] == "rent"

    def test_listing_type_synonym_renamed(self):
        df = pd.DataFrame({"city": ["Berlin"], "price": [1200], "deal_type": ["for_rent"]})
        out = DataLoaderCsv.format_df(df)
        assert "listing_type" in out.columns
        assert out.loc[0, "listing_type"] == "rent"

    def test_listing_type_sale_or_rent_normalization(self):
        df = pd.DataFrame(
            {
                "city": ["A", "B", "C", "D", "E", "F", "G"],
                "listing_type": [
                    "for_rent",
                    "rental",
                    "lease",
                    "for_sale",
                    "sold",
                    "room_rent",
                    "sublet",
                ],
                "price": [100, 200, 300, 400, 500, 600, 700],
            }
        )
        out = DataLoaderCsv.format_df(df)
        # format_df shuffles rows with random_state=1, so check by set not order
        result_set = set(out["listing_type"])
        expected_set = {"rent", "sale", "room", "sublease"}
        assert result_set == expected_set

    def test_listing_type_nan_filled_with_rent(self):
        df = pd.DataFrame(
            {"city": ["Warsaw", "Krakow"], "price": [1000, 900], "listing_type": [None, "sale"]}
        )
        out = DataLoaderCsv.format_df(df)
        assert out.loc[0, "listing_type"] == "rent"

    def test_status_synonym_renamed_to_listing_type(self):
        df = pd.DataFrame({"city": ["Berlin"], "price": [1200], "status": ["for_sale"]})
        out = DataLoaderCsv.format_df(df)
        assert "listing_type" in out.columns
        assert out.loc[0, "listing_type"] == "sale"

    def test_geo_coordinates_filled_for_known_cities(self):
        df = pd.DataFrame({"city": ["Warsaw", "Krakow"], "price": [1000, 900]})
        out = DataLoaderCsv.format_df(df)
        assert not pd.isna(out.loc[0, "latitude"])
        assert not pd.isna(out.loc[0, "longitude"])

    def test_geo_coordinates_preserved_when_present(self):
        df = pd.DataFrame(
            {"city": ["Custom"], "price": [1000], "latitude": [50.0], "longitude": [20.0]}
        )
        out = DataLoaderCsv.format_df(df)
        assert out.loc[0, "latitude"] == 50.0
        assert out.loc[0, "longitude"] == 20.0

    def test_yes_no_replaced_with_bool(self):
        df = pd.DataFrame({"city": ["A"], "price": [100], "feature": ["yes"]})
        out = DataLoaderCsv.format_df(df)
        assert out.loc[0, "feature"] is True or out.loc[0, "feature"] == True  # noqa: E712

    def test_boolean_columns_normalized(self):
        df = pd.DataFrame(
            {
                "city": ["A"],
                "price": [100],
                "has_parking": [True],
                "has_garden": [False],
                "has_pool": [True],
                "has_garage": [False],
                "has_bike_room": [True],
                "is_furnished": [False],
                "pets_allowed": [True],
                "has_balcony": [False],
                "has_elevator": [True],
            }
        )
        out = DataLoaderCsv.format_df(df)
        for col in [
            "has_parking",
            "has_garden",
            "has_pool",
            "has_garage",
            "has_bike_room",
            "is_furnished",
            "pets_allowed",
            "has_balcony",
            "has_elevator",
        ]:
            assert col in out.columns
            assert isinstance(out.loc[0, col], (bool, np.bool_))

    def test_integer_columns_converted_to_float(self):
        df = pd.DataFrame({"city": ["A"], "price": [100], "rooms": [2]})
        out = DataLoaderCsv.format_df(df)
        # rooms is int in input, should be float
        assert pd.api.types.is_float_dtype(out["rooms"]) or isinstance(out.loc[0, "rooms"], float)

    def test_bathrooms_generated_from_rooms(self):
        df = pd.DataFrame({"city": ["A", "B"], "price": [100, 200], "rooms": [2, 3]})
        out = DataLoaderCsv.format_df(df)
        assert "bathrooms" in out.columns
        for v in out["bathrooms"]:
            assert v in (1.0, 2.0)

    def test_bathrooms_existing_filled_nan(self):
        df = pd.DataFrame(
            {"city": ["A", "B"], "price": [100, 200], "rooms": [2, 3], "bathrooms": [None, 2.0]}
        )
        out = DataLoaderCsv.format_df(df)
        assert out.loc[0, "bathrooms"] == 1.0
        assert out.loc[1, "bathrooms"] == 2.0

    def test_has_garden_has_pool_generated_randomly(self):
        df = pd.DataFrame({"city": ["A"], "price": [100], "rooms": [2]})
        out = DataLoaderCsv.format_df(df)
        assert "has_garden" in out.columns
        assert "has_pool" in out.columns
        assert "has_garage" in out.columns
        assert "has_bike_room" in out.columns
        for col in ["has_garden", "has_pool", "has_garage", "has_bike_room"]:
            assert isinstance(out.loc[0, col], (bool, np.bool_))

    def test_has_elevator_generated_randomly(self):
        df = pd.DataFrame({"city": ["A"], "price": [100], "rooms": [2]})
        out = DataLoaderCsv.format_df(df)
        assert "has_elevator" in out.columns

    def test_year_built_synonym_renamed(self):
        df = pd.DataFrame({"city": ["A"], "price": [100], "construction_year": [2010]})
        out = DataLoaderCsv.format_df(df)
        assert "year_built" in out.columns

    def test_year_built_built_year_synonym(self):
        df = pd.DataFrame({"city": ["A"], "price": [100], "built_year": [2005]})
        out = DataLoaderCsv.format_df(df)
        assert "year_built" in out.columns

    def test_year_built_generated_when_missing(self):
        df = pd.DataFrame({"city": ["A"], "price": [100], "rooms": [2]})
        out = DataLoaderCsv.format_df(df)
        assert "year_built" in out.columns
        assert 1970 <= out.loc[0, "year_built"] <= 2024

    def test_year_built_nan_filled(self):
        df = pd.DataFrame({"city": ["A", "B"], "price": [100, 200], "year_built": [None, 2010]})
        out = DataLoaderCsv.format_df(df)
        assert out.loc[0, "year_built"] == 2000
        assert out.loc[1, "year_built"] == 2010

    def test_energy_rating_synonym_renamed(self):
        df = pd.DataFrame({"city": ["A"], "price": [100], "energy_class": ["B"]})
        out = DataLoaderCsv.format_df(df)
        assert "energy_rating" in out.columns

    def test_energy_rating_epc_synonym(self):
        df = pd.DataFrame({"city": ["A"], "price": [100], "epc": ["A"]})
        out = DataLoaderCsv.format_df(df)
        assert "energy_rating" in out.columns

    def test_energy_rating_generated_when_missing(self):
        df = pd.DataFrame({"city": ["A"], "price": [100], "rooms": [2]})
        out = DataLoaderCsv.format_df(df)
        assert "energy_rating" in out.columns
        assert out.loc[0, "energy_rating"] in ["A", "B", "C", "D", "E", "F", "G"]

    def test_energy_rating_nan_filled(self):
        df = pd.DataFrame({"city": ["A", "B"], "price": [100, 200], "energy_rating": [None, "A"]})
        out = DataLoaderCsv.format_df(df)
        assert out.loc[0, "energy_rating"] == "C"
        assert out.loc[1, "energy_rating"] == "A"

    def test_rows_count_limits_output(self):
        df = pd.DataFrame({"city": ["A", "B", "C"], "price": [100, 200, 300]})
        out = DataLoaderCsv.format_df(df, rows_count=2)
        assert len(out) == 2

    def test_shuffle_reproducible_with_random_state(self):
        df = pd.DataFrame({"city": ["A", "B", "C"], "price": [100, 200, 300]})
        out1 = DataLoaderCsv.format_df(df.copy())
        out2 = DataLoaderCsv.format_df(df.copy())
        # With random_state=1, should be deterministic
        assert list(out1["city"]) == list(out2["city"])

    def test_all_polish_cities_get_coordinates(self):
        """Each known Polish city in the city_coords mapping gets correct coords."""
        polish_cities = [
            "warsaw",
            "krakow",
            "wroclaw",
            "poznan",
            "gdansk",
            "szczecin",
            "lublin",
            "katowice",
            "bydgoszcz",
            "lodz",
        ]
        for city in polish_cities:
            df = pd.DataFrame({"city": [city], "price": [1000]})
            out = DataLoaderCsv.format_df(df)
            assert not pd.isna(out.loc[0, "latitude"]), f"Missing lat for {city}"
            assert not pd.isna(out.loc[0, "longitude"]), f"Missing lon for {city}"

    def test_format_df_with_empty_dataframe(self):
        df = pd.DataFrame({"city": pd.Series(dtype=str), "price": pd.Series(dtype=float)})
        out = DataLoaderCsv.format_df(df)
        assert len(out) == 0

    def test_data_not_dropped_missing_values(self):
        """format_df no longer drops rows with NaN values."""
        df = pd.DataFrame({"city": ["A", None], "price": [100, None]})
        out = DataLoaderCsv.format_df(df)
        assert len(out) == 2


# ---------------------------------------------------------------------------
# DataLoaderExcel.__init__
# ---------------------------------------------------------------------------


class TestDataLoaderExcelInit:
    """Tests for DataLoaderExcel initialization."""

    def test_init_basic(self):
        loader = DataLoaderExcel("test.xlsx")
        assert loader.sheet_name is None
        assert loader.header_row == 0
        assert loader.source_type == "excel"

    def test_init_with_custom_params(self):
        loader = DataLoaderExcel("test.xlsx", sheet_name="Sheet1", header_row=2, source_type="csv")
        assert loader.sheet_name == "Sheet1"
        assert loader.header_row == 2
        assert loader.source_type == "csv"

    def test_inherits_from_data_loader_csv(self):
        assert issubclass(DataLoaderExcel, DataLoaderCsv)


# ---------------------------------------------------------------------------
# DataLoaderExcel.get_sheet_names
# ---------------------------------------------------------------------------


class TestExcelGetSheetNames:
    """Tests for DataLoaderExcel.get_sheet_names."""

    def test_returns_empty_when_no_path(self):
        loader = DataLoaderExcel(Path("/tmp/nonexistent.xlsx"))
        loader.csv_path = None
        assert loader.get_sheet_names() == []

    def test_xlsx_sheet_names(self, monkeypatch):
        class Workbook:
            sheetnames = ["Sheet1", "Sheet2"]

            def close(self):
                return None

        import openpyxl

        monkeypatch.setattr(openpyxl, "load_workbook", lambda *_args, **_kwargs: Workbook())
        loader = DataLoaderExcel("fake.xlsx")
        assert loader.get_sheet_names() == ["Sheet1", "Sheet2"]

    def test_xls_sheet_names(self, monkeypatch):
        class Workbook:
            def sheet_names(self):
                return ["SheetA"]

        import xlrd

        monkeypatch.setattr(xlrd, "open_workbook", lambda *_args, **_kwargs: Workbook())
        loader = DataLoaderExcel("fake.xls")
        assert loader.get_sheet_names() == ["SheetA"]

    def test_ods_primary_path(self, monkeypatch):
        from types import SimpleNamespace

        import odf.opendocument

        doc = SimpleNamespace(spreadsheets={"ODSMain": object()})
        monkeypatch.setattr(odf.opendocument, "load", lambda *_args, **_kwargs: doc)
        loader = DataLoaderExcel("fake.ods")
        assert loader.get_sheet_names() == ["ODSMain"]

    def test_ods_fallback_path(self, monkeypatch):
        class ExcelFile:
            sheet_names = ["ODS1"]

            def close(self):
                return None

        import odf.opendocument

        monkeypatch.setattr(
            odf.opendocument,
            "load",
            lambda *_args, **_kwargs: (_ for _ in ()).throw(ImportError()),
        )
        monkeypatch.setattr(pd, "ExcelFile", lambda *_args, **_kwargs: ExcelFile())
        loader = DataLoaderExcel("fake.ods")
        assert loader.get_sheet_names() == ["ODS1"]

    def test_other_suffix_uses_pandas(self, monkeypatch):
        class ExcelFile:
            sheet_names = ["MacroSheet"]

            def close(self):
                return None

        monkeypatch.setattr(pd, "ExcelFile", lambda *_args, **_kwargs: ExcelFile())
        loader = DataLoaderExcel("fake.xlsm")
        assert loader.get_sheet_names() == ["MacroSheet"]

    def test_import_error_raises(self, monkeypatch):
        import openpyxl

        monkeypatch.setattr(
            openpyxl,
            "load_workbook",
            lambda *_args, **_kwargs: (_ for _ in ()).throw(ImportError("no openpyxl")),
        )
        loader = DataLoaderExcel("fake.xlsx")
        with pytest.raises(ImportError, match="Excel file reading requires"):
            loader.get_sheet_names()


# ---------------------------------------------------------------------------
# DataLoaderExcel.load_df
# ---------------------------------------------------------------------------


class TestExcelLoadDf:
    """Tests for DataLoaderExcel.load_df."""

    def test_no_path_raises_value_error(self):
        loader = DataLoaderExcel("fake.xlsx")
        loader.csv_path = None
        with pytest.raises(ValueError, match="No Excel file path provided"):
            loader.load_df()

    def test_non_excel_suffix_raises_value_error(self):
        loader = DataLoaderExcel("fake.csv")
        with pytest.raises(ValueError, match="Unsupported file format"):
            loader.load_df()

    def test_xlsx_loads_with_openpyxl_engine(self, monkeypatch):
        captured = {}

        def fake_read_excel(_path, **kwargs):
            captured.update(kwargs)
            return pd.DataFrame({"city": ["Wroclaw"], "price": [1000], "rooms": [2]})

        monkeypatch.setattr(pd, "read_excel", fake_read_excel)
        loader = DataLoaderExcel("fake.xlsx", sheet_name="Sheet1", header_row=1)
        df_out = loader.load_df()
        assert captured["engine"] == "openpyxl"
        assert captured["sheet_name"] == "Sheet1"
        assert captured["header"] == 1
        assert len(df_out) == 1

    def test_xls_loads_with_xlrd_engine(self, monkeypatch):
        captured = {}

        def fake_read_excel(_path, **kwargs):
            captured.update(kwargs)
            return pd.DataFrame({"a": [1]})

        monkeypatch.setattr(pd, "read_excel", fake_read_excel)
        loader = DataLoaderExcel("fake.xls")
        loader.load_df()
        assert "engine" in captured
        assert captured["engine"] == "xlrd"

    def test_ods_loads_with_odf_engine(self, monkeypatch):
        import sys

        captured = {}

        def fake_read_excel(_path, **kwargs):
            captured.update(kwargs)
            return pd.DataFrame({"a": [1]})

        # Fix the odf stub module to have a valid __spec__ so find_spec doesn't crash
        if "odf" in sys.modules:
            from importlib.machinery import ModuleSpec

            sys.modules["odf"].__spec__ = ModuleSpec("odf", None)

        monkeypatch.setattr(pd, "read_excel", fake_read_excel)
        loader = DataLoaderExcel("fake.ods")
        loader.load_df()
        assert "engine" in captured
        assert captured["engine"] == "odf"

    def test_no_sheet_name_no_header_in_kwargs(self, monkeypatch):
        captured = {}

        def fake_read_excel(_path, **kwargs):
            captured.update(kwargs)
            return pd.DataFrame({"a": [1]})

        monkeypatch.setattr(pd, "read_excel", fake_read_excel)
        loader = DataLoaderExcel("fake.xlsx")
        loader.load_df()
        assert "sheet_name" not in captured
        assert captured["header"] == 0  # header_row defaults to 0

    def test_import_error_raises(self, monkeypatch):
        monkeypatch.setattr(
            pd, "read_excel", lambda *_args, **_kwargs: (_ for _ in ()).throw(ImportError("nope"))
        )
        loader = DataLoaderExcel("fake.xlsx")
        with pytest.raises(ImportError, match="Excel input requires optional"):
            loader.load_df()

    def test_generic_exception_raises(self, monkeypatch):
        monkeypatch.setattr(
            pd, "read_excel", lambda *_args, **_kwargs: (_ for _ in ()).throw(Exception("corrupt"))
        )
        loader = DataLoaderExcel("fake.xlsx")
        with pytest.raises(Exception, match="Failed to load Excel file"):
            loader.load_df()

    def test_load_df_logs_info(self, caplog):
        """load_df logs info about loaded data."""
        with tempfile.TemporaryDirectory() as tmp:
            p = Path(tmp) / "data.xlsx"
            pd.DataFrame({"city": ["Test"]}).to_excel(p, index=False)
            loader = DataLoaderExcel(str(p))
            with caplog.at_level("INFO"):
                loader.load_df()
            assert "Excel data loaded" in caplog.text


# ---------------------------------------------------------------------------
# DataLoaderExcel.detect_source_type
# ---------------------------------------------------------------------------


class TestDetectSourceType:
    """Tests for DataLoaderExcel.detect_source_type."""

    def test_csv_file(self):
        assert DataLoaderExcel.detect_source_type("data.csv") == "csv"

    def test_xlsx_file(self):
        assert DataLoaderExcel.detect_source_type("data.xlsx") == "excel"

    def test_xls_file(self):
        assert DataLoaderExcel.detect_source_type("data.xls") == "excel"

    def test_ods_file(self):
        assert DataLoaderExcel.detect_source_type("data.ods") == "excel"

    def test_unknown_extension(self):
        assert DataLoaderExcel.detect_source_type("data.json") == "unknown"

    def test_no_extension(self):
        assert DataLoaderExcel.detect_source_type("datafile") == "unknown"

    def test_http_url(self):
        assert DataLoaderExcel.detect_source_type("http://example.com/data.csv") == "url"

    def test_https_url(self):
        assert DataLoaderExcel.detect_source_type("https://example.com/data.xlsx") == "url"

    def test_url_instance(self):
        url = URL("https://example.com/data.xlsx")
        assert DataLoaderExcel.detect_source_type(url) == "url"

    def test_github_url_detected_as_url(self):
        url = "https://github.com/user/repo/blob/main/data.csv"
        assert DataLoaderExcel.detect_source_type(url) == "url"

    def test_path_object_csv(self):
        assert DataLoaderExcel.detect_source_type(Path("data.csv")) == "csv"

    def test_path_object_excel(self):
        assert DataLoaderExcel.detect_source_type(Path("data.xlsx")) == "excel"
